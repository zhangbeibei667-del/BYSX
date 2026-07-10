from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - runtime dependency hint
    raise SystemExit(
        "缺少 openpyxl。请先安装：python -m pip install openpyxl，"
        "或使用 Codex bundled Python 运行本脚本。"
    ) from exc


SOURCE = "SymMap v2.0"


FILE_SPECS = {
    "SMTS": {
        "glob": "*SMTS*file*.xlsx",
        "type": "症状",
        "id_col": "TCM_symptom_id",
        "name_col": "TCM_symptom_name",
        "desc_col": "Symptom_definition",
        "fields": {
            "pinyin": "Symptom_pinYin",
            "locus": "Symptom_locus",
            "property": "Symptom_property",
            "symmap_type": "Type",
            "version": "Version",
        },
    },
    "SMSY": {
        "glob": "*SMSY*file*.xlsx",
        "type": "证候",
        "id_col": "Syndrome_id",
        "name_col": "Syndrome_name",
        "desc_col": "Syndrome_definition",
        "fields": {
            "english": "Syndrome_English",
            "pinyin": "Syndrome_PinYin",
            "symmap_type": "Type",
            "version": "Version",
        },
    },
    "SMHB": {
        "glob": "*SMHB*file*.xlsx",
        "type": "药材",
        "id_col": "Herb_id",
        "name_col": "Chinese_name",
        "desc_col": "",
        "alias_col": "Alias",
        "fields": {
            "pinyin": "Pinyin_name",
            "latin": "Latin_name",
            "english": "English_name",
            "properties_chinese": "Properties_Chinese",
            "meridians_chinese": "Meridians_Chinese",
            "class_chinese": "Class_Chinese",
            "use_part": "UsePart",
            "tcmid_id": "TCMID_id",
            "tcm_id_id": "TCM-ID_id",
            "tcmsp_id": "TCMSP_id",
            "herbdb_id": "HERBDB_ID",
        },
    },
    "SMDE": {
        "glob": "*SMDE*file*.xlsx",
        "type": "疾病",
        "id_col": "Disease_id",
        "name_col": "Disease_Name",
        "desc_col": "Disease_definition",
        "fields": {
            "umls_id": "UMLS_id",
            "mesh_id": "MeSH_id",
            "omim_id": "OMIM_id",
            "orphanet_id": "Orphanet_id",
            "icd10cm_id": "ICD10CM_id",
            "meddra_id": "MedDRA_id",
            "version": "Version",
            "link_disease_id": "Link_disease_id",
        },
    },
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"none", "nan", "null"}:
        return ""
    return text


def safe_id(prefix: str, raw_id: str, fallback_index: int) -> str:
    raw = clean(raw_id)
    if not raw:
        return f"{prefix}{fallback_index:05d}"
    raw = re.sub(r"\.0$", "", raw)
    if raw.upper().startswith(prefix):
        return raw.upper()
    digits = re.sub(r"\D", "", raw)
    if digits:
        return f"{prefix}{int(digits):05d}"
    slug = re.sub(r"[^0-9A-Za-z_]+", "_", raw).strip("_")
    return f"{prefix}_{slug}" if slug else f"{prefix}{fallback_index:05d}"


def is_suppressed(row: dict[str, Any]) -> bool:
    value = clean(row.get("Suppress")).lower()
    return value in {"1", "true", "yes", "y", "suppressed"}


def row_dicts(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    result: list[dict[str, Any]] = []
    for values in rows:
        result.append({headers[index]: values[index] if index < len(values) else None for index in range(len(headers))})
    workbook.close()
    return result


def find_file(input_dir: Path, pattern: str) -> Path:
    matches = sorted(input_dir.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"未找到匹配文件：{input_dir / pattern}")
    return matches[0]


def load_existing_names(entities_dir: Path) -> set[tuple[str, str]]:
    existing: set[tuple[str, str]] = set()
    for path in entities_dir.glob("*.json"):
        if path.name == "entities_symmap_import.json":
            continue
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(data, list):
            continue
        for item in data:
            name = clean(item.get("name"))
            entity_type = clean(item.get("type"))
            if name and entity_type:
                existing.add((entity_type, name))
    return existing


def build_description(spec_key: str, row: dict[str, Any], desc_col: str) -> str:
    desc = clean(row.get(desc_col)) if desc_col else ""
    if spec_key == "SMHB":
        parts = [
            clean(row.get("Properties_Chinese")),
            clean(row.get("Meridians_Chinese")),
            clean(row.get("Class_Chinese")),
            clean(row.get("UsePart")),
        ]
        desc = "；".join(part for part in parts if part)
    return desc


def import_entities(input_dir: Path, output_dir: Path) -> dict[str, int]:
    entities_dir = output_dir / "entities"
    entities_dir.mkdir(parents=True, exist_ok=True)
    existing = load_existing_names(entities_dir)

    imported: list[dict[str, Any]] = []
    stats = {"read": 0, "imported": 0, "duplicates": 0, "suppressed": 0}

    for spec_key, spec in FILE_SPECS.items():
        path = find_file(input_dir, spec["glob"])
        rows = row_dicts(path)
        for index, row in enumerate(rows, start=1):
            stats["read"] += 1
            if is_suppressed(row):
                stats["suppressed"] += 1
                continue

            name = clean(row.get(spec["name_col"]))
            entity_type = spec["type"]
            if not name:
                continue
            if (entity_type, name) in existing:
                stats["duplicates"] += 1
                continue

            properties: dict[str, Any] = {
                "source": SOURCE,
                "symmap_component": spec_key,
                "symmap_id": clean(row.get(spec["id_col"])),
            }
            for prop_key, column in spec["fields"].items():
                value = clean(row.get(column))
                if value:
                    properties[prop_key] = value

            entity = {
                "id": safe_id(spec_key, clean(row.get(spec["id_col"])), index),
                "name": name,
                "type": entity_type,
                "alias": clean(row.get(spec.get("alias_col", ""))),
                "description": build_description(spec_key, row, spec.get("desc_col", "")),
                "properties": properties,
            }
            imported.append(entity)
            existing.add((entity_type, name))
            stats["imported"] += 1

    output_path = entities_dir / "entities_symmap_import.json"
    output_path.write_text(
        json.dumps(imported, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    stats["output_file"] = str(output_path)
    return stats


def main() -> None:
    parser = argparse.ArgumentParser(description="Import SymMap v2 entity tables into the local GraphRAG KG.")
    parser.add_argument("--input-dir", required=True, help="Directory containing the downloaded SymMap v2 xlsx files.")
    parser.add_argument(
        "--output-dir",
        required=True,
        help="GraphRAG data directory, e.g. integrated_entities_graphrag/data",
    )
    args = parser.parse_args()

    stats = import_entities(Path(args.input_dir), Path(args.output_dir))
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
