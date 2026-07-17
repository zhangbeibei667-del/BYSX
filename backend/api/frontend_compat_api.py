from __future__ import annotations

import json
import csv
import io
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Literal

from fastapi import APIRouter, Body, File, HTTPException, Query, UploadFile
from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

from backend.services.agent_service import AgentService
from backend.services.conversation_service import ConversationService
from backend.services.history_service import HistoryService
from backend.services.rag_service import RAGService
from backend.services.local_graphrag_service import Entity, Relation, get_local_graphrag_service
from backend.services.teaching_case_service import TeachingCaseService
from backend.services.llm_client import get_llm_client
from backend.services.document_service import DocumentService
from backend.tools.rag_tool import RAGRetrievalTool


router = APIRouter(prefix="/api", tags=["Frontend compatibility"])

class ChatAskRequest(BaseModel):
    question: str
    historyId: str | None = None
    answerMode: Literal["concise", "teaching", "deep"] = "concise"


class CaseCreateRequest(BaseModel):
    name: str | None = ""
    age: int | None = None
    gender: str | None = ""
    chiefComplaint: str | None = ""
    tongueExam: str | None = ""
    pulseExam: str | None = ""
    teachingNote: str | None = ""
    analysisResult: dict[str, Any] | None = None


ENTITY_TYPE_BY_ADMIN_KIND = {
    "herbs": ("药材", "H"),
    "prescriptions": ("方剂", "F"),
    "symptoms": ("症状", "S"),
    "syndromes": ("证候", "Z"),
}


def graph_data_root() -> Path:
    return Path(__file__).resolve().parents[2] / "data"


def graph_repo_root() -> Path:
    return Path(__file__).resolve().parents[2]


def admin_entities_path() -> Path:
    return graph_repo_root() / "entities" / "entities_admin_supplement.json"


def admin_relations_path() -> Path:
    return graph_repo_root() / "relations" / "relations_admin_supplement.json"


def admin_overlay_path() -> Path:
    return graph_data_root() / "admin_overlay.json"


def read_json_file(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        return default


def write_json_file(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def read_admin_entities() -> list[dict[str, Any]]:
    data = read_json_file(admin_entities_path(), {"entities": []})
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    return [item for item in data.get("entities", []) if isinstance(item, dict)]


def write_admin_entities(items: list[dict[str, Any]]) -> None:
    write_json_file(admin_entities_path(), {"entities": items})
    get_local_graphrag_service.cache_clear()


def read_admin_relations() -> list[dict[str, Any]]:
    data = read_json_file(admin_relations_path(), {"relations": []})
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    return [item for item in data.get("relations", []) if isinstance(item, dict)]


def write_admin_relations(items: list[dict[str, Any]]) -> None:
    write_json_file(admin_relations_path(), {"relations": items})
    get_local_graphrag_service.cache_clear()


def read_admin_overlay() -> dict[str, list[str]]:
    data = read_json_file(admin_overlay_path(), {})
    return {
        "deleted_entity_ids": [str(item) for item in data.get("deleted_entity_ids", [])],
        "deleted_relation_keys": [str(item) for item in data.get("deleted_relation_keys", [])],
    }


def write_admin_overlay(data: dict[str, list[str]]) -> None:
    write_json_file(admin_overlay_path(), data)
    get_local_graphrag_service.cache_clear()


def next_entity_id(entity_type: str, prefix: str) -> str:
    service = get_local_graphrag_service()
    max_num = 0
    for entity in service.entities:
        if entity.type != entity_type or not entity.id.upper().startswith(prefix):
            continue
        suffix = entity.id[len(prefix) :]
        if suffix.isdigit():
            max_num = max(max_num, int(suffix))
    return f"{prefix}{max_num + 1:04d}"


def entity_payload_to_item(payload: dict[str, Any], entity_type: str, prefix: str, entity_id: str | None = None) -> dict[str, Any]:
    props = payload.get("properties") if isinstance(payload.get("properties"), dict) else {}
    category = payload.get("category") or props.get("category") or entity_type
    props = {**props, "category": category, "source": props.get("source") or "后台管理录入"}
    return {
        "id": entity_id or str(payload.get("id") or next_entity_id(entity_type, prefix)),
        "name": str(payload.get("name") or "").strip(),
        "type": entity_type,
        "alias": str(payload.get("alias") or "").strip(),
        "description": str(payload.get("description") or "").strip(),
        "properties": props,
    }


def relation_key_from_parts(source_id: str, relation: str, target_id: str) -> str:
    return f"{source_id}|{relation}|{target_id}"


def relation_to_admin_item(relation: Relation | dict[str, Any]) -> dict[str, Any]:
    if isinstance(relation, Relation):
        source_id = relation.source_id
        rel = relation.relation
        target_id = relation.target_id
        item = {
            "source_id": source_id,
            "source_name": relation.source_name,
            "relation": rel,
            "target_id": target_id,
            "target_name": relation.target_name,
            "evidence": relation.evidence,
        }
    else:
        source_id = str(relation.get("source_id", ""))
        rel = str(relation.get("relation", ""))
        target_id = str(relation.get("target_id", ""))
        item = dict(relation)
    item["id"] = relation_key_from_parts(source_id, rel, target_id)
    return item


def ok(data: Any) -> dict[str, Any]:
    return {"code": 200, "msg": "ok", "data": data}


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def build_relation_property_index(service) -> dict[str, dict[str, list[str]]]:
    """把规范化图谱关系聚合成管理表单需要的可读属性。"""
    index: dict[str, dict[str, list[str]]] = {}
    for relation in service.relations:
        source = service.entities_by_id.get(relation.source_id)
        target = service.entities_by_id.get(relation.target_id)
        if source is None or target is None:
            continue
        field = ""
        if relation.relation == "具有" and target.type == "性味":
            field = "nature_and_flavor"
        elif relation.relation == "具有" and target.type == "功效":
            field = "functions" if source.type == "方剂" else "efficacy"
        elif relation.relation == "归经" and target.type == "归经":
            field = "channel_tropism"
        elif relation.relation == "禁忌" and target.type == "禁忌":
            field = "contraindications"
        elif relation.relation == "包含" and source.type == "方剂" and target.type == "药材":
            field = "composition"
        elif relation.relation == "主治" and target.type in {"疾病", "症状", "证候"}:
            field = "indications"
        elif relation.relation == "提示" and source.type == "症状" and target.type == "证候":
            field = "differentiation"
        if field:
            values = index.setdefault(relation.source_id, {}).setdefault(field, [])
            if target.name not in values:
                values.append(target.name)

        # 为证候管理页生成反向可读字段，不改变图谱关系方向。
        if relation.relation == "提示" and source.type == "症状" and target.type == "证候":
            values = index.setdefault(target.id, {}).setdefault("clinical_manifestations", [])
            if source.name not in values:
                values.append(source.name)
        if relation.relation == "主治" and source.type == "方剂" and target.type == "证候":
            values = index.setdefault(target.id, {}).setdefault("commonly_used_formulas", [])
            if source.name not in values:
                values.append(source.name)
    return index


def entity_to_admin(
    entity: Entity,
    relation_properties: dict[str, dict[str, list[str]]] | None = None,
) -> dict[str, Any]:
    props = dict(entity.properties or {})
    for field, values in (relation_properties or {}).get(entity.id, {}).items():
        if not props.get(field) and values:
            props[field] = "、".join(values)

    # 药材原始 description 中包含“性味。归经。分类药。”，归经尚未作为
    # 药材关系单独保存时从原始说明提取，避免管理页丢失已有数据。
    derived_category = ""
    if entity.type == "药材" and entity.description:
        sections = [part.strip() for part in entity.description.split("。") if part.strip()]
        if sections and not props.get("nature_and_flavor"):
            props["nature_and_flavor"] = sections[0].replace(",", "、").replace("，", "、")
        meridian_match = re.search(r"归([^。]+?经)(?:。|$)", entity.description)
        if meridian_match and not props.get("channel_tropism"):
            props["channel_tropism"] = meridian_match.group(1).replace(",", "、").replace("，", "、")
        if len(sections) >= 3 and sections[-1].endswith("药"):
            derived_category = sections[-1]

    return {
        "id": entity.id,
        "name": entity.name,
        "type": entity.type,
        "alias": entity.alias,
        "category": props.get("category") or derived_category or entity.type,
        "description": entity.description,
        "properties": props,
        "createdAt": props.get("createdAt") or "",
        "updatedAt": props.get("updatedAt") or "",
    }


def graph_response(nodes: list[dict], edges: list[dict]) -> dict[str, Any]:
    return {
        "nodes": nodes,
        "edges": edges,
    }


def case_text_from_payload(payload: CaseCreateRequest) -> str:
    parts = [
        payload.chiefComplaint or "",
        payload.tongueExam or "",
        payload.pulseExam or "",
    ]
    return "，".join(part.strip("，。 \n\r\t") for part in parts if part and part.strip())


def is_case_like_question(text: str) -> bool:
    """Route real case narratives to the Agent, and knowledge questions to GraphRAG."""
    stripped = text.strip()
    if not stripped:
        return False

    case_markers = ["患者", "近", "主诉", "舌", "苔", "脉", "夜间", "大便", "小便", "盗汗", "纳差"]
    question_markers = [
        "是什么",
        "有哪些",
        "有什么",
        "为什么",
        "由哪些",
        "包含哪些",
        "组成",
        "功效",
        "禁忌",
        "解释",
        "介绍",
    ]

    marker_count = sum(1 for marker in case_markers if marker in stripped)
    asks_knowledge = any(marker in stripped for marker in question_markers)

    if stripped.startswith(("患者", "病例", "男", "女")):
        return True
    if "请进行中医" in stripped or "辨证" in stripped:
        return True
    if marker_count >= 3 and not asks_knowledge:
        return True
    return False


def evidence_to_sources(evidence: list[dict[str, Any]], *, clinical: bool = False) -> list[dict[str, Any]]:
    """Expose provenance separately from the natural-language answer body."""
    sources = []
    for item in evidence:
        locator = item.get("locator") if isinstance(item.get("locator"), dict) else {}
        is_graph_relation = (
            str(item.get("source") or "") == "知识图谱"
            or str(item.get("content") or "").startswith("知识图谱关系：")
        )
        is_document = bool(item.get("category") or item.get("document_id"))
        is_graph_entity = not is_document and not is_graph_relation
        raw_score = item.get("score")
        normalized_score = None
        metric_label = None
        status_label = None
        if is_document and raw_score is not None:
            normalized_score = max(0.0, min(1.0, float(raw_score)))
            metric_label = "检索相关度"
        elif is_graph_relation:
            status_label = "已验证关系"
        elif is_graph_entity:
            status_label = "命中图谱实体"
        original_text = str(item.get("content") or item.get("excerpt") or "")
        contains_treatment = any(marker in original_text for marker in ("治宜", "方用", "方选", "药用"))
        if clinical and is_document and contains_treatment:
            original_text = RAGRetrievalTool._sanitize_clinical_evidence(original_text)
        sources.append({
            "title": item.get("title") or item.get("citation") or "未命名资料",
            "type": (
                "图谱关系" if is_graph_relation else "图谱实体" if is_graph_entity
                else item.get("source_type") or item.get("category") or item.get("type") or "文献"
            ),
            "source_detail": item.get("citation") or item.get("source") or "",
            "original_text": original_text,
            "chapter": item.get("chapter") or locator.get("chapter") or locator.get("page") or "",
            "score": normalized_score,
            "metric_label": metric_label,
            "status_label": status_label,
            "contains_treatment": contains_treatment,
        })
    return sources


def answer_with_graphrag(question: str, *, generate_answer: bool = True,
                         answer_mode: str = "concise") -> dict[str, Any]:
    result = RAGService().search(
        question, top_k=5, generate_answer=generate_answer, answer_mode=answer_mode
    )
    evidence = result.get("evidence", [])
    return {
        "query": result.get("query", question),
        "answer": result.get("answer", ""),
        "symptoms": [],
        "tongue": [],
        "pulse": [],
        "syndromes": result.get("syndromes", []),
        "formulas": result.get("formulas", []),
        "herbs": result.get("herbs", []),
        "follow_up_questions": result.get("follow_up_questions", []),
        "needs_clarification": result.get("needs_clarification", False),
        "graph": result.get("graph", {"nodes": [], "edges": []}),
        "evidence": evidence,
        "sources": evidence_to_sources(
            evidence, clinical=RAGRetrievalTool.is_personal_clinical_request(question)
        ),
        "intent": result.get("intent"),
        "intent_label": result.get("intent_label"),
        "mode": result.get("mode", "local-graphrag"),
        "citations": result.get("citations", []),
        "retrieval": result.get("retrieval", {}),
        "generation": result.get("generation", {"mode": "local-evidence-template"}),
        "evidence_confidence": result.get("evidence_confidence", "insufficient"),
        "answer_mode": result.get("answer_mode", answer_mode),
        "clarification_topic": result.get("clarification_topic"),
        "clarification_progress": result.get("clarification_progress"),
        "clinical_dimensions": result.get("clinical_dimensions", {}),
        "differential_evidence": result.get("differential_evidence", []),
        "safety_notice": "内容用于中医药知识学习与辨证辅助，不替代医师诊断和处方。",
    }


def attach_teaching_agent_trace(result: dict[str, Any]) -> dict[str, Any]:
    """Expose the real graph/RAG teaching workflow instead of presenting it as a black box."""
    retrieval = result.get("retrieval", {})
    document_hits = int(retrieval.get("document_hits") or 0)
    graph_hits = int(retrieval.get("graph_hits") or 0)
    needs_clarification = bool(result.get("needs_clarification"))
    steps = [
        {
            "name": "症状追问 Agent",
            "status": "awaiting_input" if needs_clarification else "completed",
            "summary": (
                f"仍需补充 {len(result.get('follow_up_questions', []))} 组辨证信息"
                if needs_clarification else "已整理本轮症状、病程、兼症与舌脉信息"
            ),
        },
        {
            "name": "知识图谱查询",
            "status": "completed",
            "summary": f"召回 {graph_hits} 条症状—证候关系证据",
        },
        {
            "name": "RAG 文献检索",
            "status": "completed" if document_hits else "insufficient",
            "summary": f"召回 {document_hits} 条可定位知识库片段",
        },
        {
            "name": "知识解释与安全审查",
            "status": "completed",
            "summary": "融合图谱与文献进行教学性鉴别，并阻止个体化处方输出",
        },
    ]
    result["agent_plan"] = {
        "agent": "诊疗教学 Agent",
        "tools": ["symptom_followup", "graph_query", "literature_search", "knowledge_explanation", "safety_review"],
        "evidence_fusion": "knowledge-graph+rag",
    }
    result["agent_steps"] = steps
    result["evidence_summary"] = {
        "graph_relations": graph_hits,
        "knowledge_base_documents": document_hits,
        "sources_available": len(result.get("sources", [])),
    }
    return result


@router.post("/case/create")
def create_case(payload: CaseCreateRequest) -> dict[str, Any]:
    return ok(TeachingCaseService().create(payload.model_dump()))


@router.post("/case/analyze/{case_id}")
def analyze_case(case_id: str) -> dict[str, Any]:
    cases = TeachingCaseService()
    case_data = cases.get(case_id)
    if not case_data:
        raise HTTPException(status_code=404, detail="case not found")
    text = "，".join(
        part.strip("，。 \n\r\t")
        for part in [
            str(case_data.get("chiefComplaint") or ""),
            str(case_data.get("tongueExam") or ""),
            str(case_data.get("pulseExam") or ""),
        ]
        if part and part.strip()
    )
    session_id = case_data.get("sessionId") or f"case-{case_id}"
    conversations = ConversationService()
    session = conversations.load(session_id)
    contextualized = conversations.contextualize(session, text)
    result = AgentService().analyze_case(contextualized, has_context=bool(session["turns"]))
    conversations.save_turn(session, text, result)
    result["conversation"] = {"id": session_id, "status": session["status"],
                              "collected": session["collected"],
                              "pending_questions": session["pending_questions"]}
    cases.save_analysis(case_id, result, session_id)
    return ok(result)


@router.get("/case/list")
def list_cases(page: int = 1, pageSize: int = 20) -> dict[str, Any]:
    return ok(TeachingCaseService().list(page, pageSize))


@router.get("/case/{case_id}")
def get_case(case_id: str) -> dict[str, Any]:
    case_data = TeachingCaseService().get(case_id)
    if not case_data:
        raise HTTPException(status_code=404, detail="case not found")
    return ok(case_data)


@router.delete("/case/{case_id}")
def delete_case(case_id: str) -> dict[str, Any]:
    return ok({"deleted": TeachingCaseService().delete(case_id)})


@router.post("/chat/ask")
def ask_chat(payload: ChatAskRequest) -> dict[str, Any]:
    return prepare_chat(payload, generate_answer=True)


def prepare_chat(payload: ChatAskRequest, *, generate_answer: bool) -> dict[str, Any]:
    conversations = ConversationService()
    session = conversations.load(payload.historyId)
    contextualized = conversations.contextualize(session, payload.question)
    clinical_query = contextualized
    last_result = session.get("last_result", {})
    awaiting_clarification = bool(session["turns"]) and session.get("status") == "awaiting_clarification"
    resumable_clarification = bool(session["turns"]) and (
        last_result.get("mode") == "evidence-insufficient"
        and bool(last_result.get("clarification_progress"))
        and len(payload.question.strip()) <= 100
        and not payload.question.rstrip().endswith(("?", "？"))
    )
    has_clarification_context = awaiting_clarification or resumable_clarification
    is_reply = has_clarification_context and (
        not payload.question.rstrip().endswith(("?", "？"))
        or ConversationService.requires_context(payload.question)
    )
    continuing_case = has_clarification_context and is_reply
    previous_result = session.get("last_result", {}) if continuing_case else {}
    previous_is_gated_rag = (
        previous_result.get("mode") == "evidence-needs-clarification"
        or (
            previous_result.get("mode") == "evidence-insufficient"
            and bool(previous_result.get("clarification_progress") or previous_result.get("clarification_topic"))
        )
    )
    if continuing_case and previous_is_gated_rag:
        topic_query = str(previous_result.get("clarification_topic") or previous_result.get("query") or "")
        prior_user_turns = [
            str(turn.get("content") or "") for turn in session["turns"] if turn.get("role") == "user"
        ]
        # The first turn is the question that triggered clarification, not an answer to it.
        prior_user_text = "\n".join(prior_user_turns[1:])
        clinical_query = (
            f"{topic_query}\n补充信息：\n{prior_user_text}\n{payload.question}"
        ).strip()
        progress = RAGRetrievalTool.clarification_progress(
            topic_query, f"{prior_user_text}\n{payload.question}"
        )
        if progress["complete"]:
            result = answer_with_graphrag(
                clinical_query, generate_answer=generate_answer, answer_mode=payload.answerMode
            )
            result["clarification_progress"] = progress
            result["clinical_dimensions"] = progress["clinical_dimensions"]
            result["differential_evidence"] = RAGRetrievalTool.build_differential_evidence(
                clinical_query, result
            )
        elif progress["finished"]:
            result = {
                **previous_result,
                "mode": "evidence-insufficient",
                "answer": (
                    "已记录你能提供的信息。由于关键舌脉信息目前无法确认，现有内容仍不足以可靠判断具体证候。"
                    "系统不会继续猜测主证，也不会据此推荐方剂、药材、剂量、穴位或食疗。"
                    "可以继续保持规律作息，减少睡前咖啡因、酒精和电子屏幕刺激；"
                    "如果睡眠问题持续存在或明显影响白天功能，建议由专业人员面诊评估。"
                ),
                "follow_up_questions": [],
                "needs_clarification": False,
                "formulas": [],
                "herbs": [],
                "generation": {"mode": "evidence-gated-insufficient", "calls": 0},
                "evidence_confidence": "insufficient",
                "clarification_progress": progress,
                "clinical_dimensions": progress["clinical_dimensions"],
                "differential_evidence": RAGRetrievalTool.build_differential_evidence(
                    clinical_query, previous_result
                ),
            }
        else:
            answered_count = len(progress["answered"])
            result = {
                **previous_result,
                "answer": (
                    f"已记录你这次补充的信息（已完成 {answered_count} 项）。"
                    "目前仍不足以可靠判断具体证候，也不会据此推荐方剂、药材或穴位。"
                    "请继续补充下面尚缺的信息；如果舌象或脉象不清楚，可以直接说明“不清楚”。"
                ),
                "follow_up_questions": progress["remaining_questions"],
                "needs_clarification": True,
                "formulas": [],
                "herbs": [],
                "generation": {"mode": "evidence-gated-clarification", "calls": 0},
                "clarification_progress": progress,
                "clinical_dimensions": progress["clinical_dimensions"],
                "differential_evidence": RAGRetrievalTool.build_differential_evidence(
                    clinical_query, previous_result
                ),
            }
    elif is_case_like_question(payload.question) or continuing_case:
        result = AgentService().analyze_case(
            contextualized,
            has_context=bool(session["turns"]),
            generate_answer=generate_answer,
        )
    else:
        rag_question = (
            contextualized
            if ConversationService.requires_context(payload.question)
            else payload.question
        )
        result = answer_with_graphrag(
            rag_question, generate_answer=generate_answer, answer_mode=payload.answerMode
        )
    clinical_context = (
        continuing_case
        or is_case_like_question(payload.question)
        or RAGRetrievalTool.is_personal_clinical_request(payload.question)
    )
    generation_mode = str(result.get("generation", {}).get("mode") or "")
    already_gated = generation_mode.startswith("evidence-gated") or generation_mode == "retrieval-only"
    if clinical_context and generation_mode.startswith("llm"):
        result["answer"] = RAGRetrievalTool.sanitize_clinical_answer(
            str(result.get("answer") or "")
        )
    if clinical_context and not already_gated and not RAGRetrievalTool.clinical_output_is_safe(
        str(result.get("answer") or "")
    ):
        fallback_answer = RAGRetrievalTool.clinical_safety_fallback(clinical_query, result)
        result = {
            **result,
            "answer": fallback_answer,
            "formulas": [],
            "herbs": [],
            "generation": {"mode": "agent-clinical-safety-fallback", "calls": 1},
            "evidence_confidence": "insufficient",
            "clinical_dimensions": result.get("clinical_dimensions", {}),
            "differential_evidence": result.get("differential_evidence", []),
        }
    if clinical_context:
        result.setdefault("sources", evidence_to_sources(result.get("evidence", []), clinical=True))
        attach_teaching_agent_trace(result)
    conversations.save_turn(session, payload.question, result)
    history_id = session["id"]
    result["conversation"] = {
        "id": history_id, "status": session["status"], "turn_count": len(session["turns"]) // 2,
        "collected": session["collected"], "pending_questions": session["pending_questions"],
    }
    if continuing_case:
        result["followup_transition"] = {
            "type": "clarification_applied",
            "answered_questions": previous_result.get("follow_up_questions", []),
            "previous_syndromes": previous_result.get("syndromes", []),
            "current_syndromes": result.get("syndromes", []),
            "previous_formulas": previous_result.get("formulas", []),
            "current_formulas": result.get("formulas", []),
        }
    return result


@router.get("/chat/ask/stream")
def ask_chat_token_stream(
    question: str,
    historyId: str | None = None,
    answerMode: Literal["concise", "teaching", "deep"] = "concise",
) -> StreamingResponse:
    result = prepare_chat(
        ChatAskRequest(question=question, historyId=historyId, answerMode=answerMode),
        generate_answer=False,
    )
    fallback_text = result.get("answer", "")
    session_id = result.get("conversation", {}).get("id")

    def events():
        import re
        client = get_llm_client()
        streamed: list[str] = []
        llm_generated = False
        is_retrieval_only = result.get("generation", {}).get("mode") == "retrieval-only"
        can_generate = is_retrieval_only and result.get("mode") not in {
            "evidence-insufficient", "evidence-needs-clarification", "retrieval-error"
        }
        messages = RAGRetrievalTool.build_generation_messages(
            str(result.get("query") or question), result, answerMode
        )
        query_for_safety = str(result.get("query") or question)
        clinical_request = RAGRetrievalTool.is_personal_clinical_request(query_for_safety)
        generation_mode = ""
        generation_usage: dict[str, Any] = {}
        if client.available and can_generate and clinical_request:
            # Clinical answers use the provider's real token stream, but are only released
            # after a complete sentence has passed the safety gate.  This keeps the UI and
            # speech output genuinely progressive without exposing an unchecked half-sentence.
            result["formulas"] = []
            result["herbs"] = []
            sentence_buffer = ""
            chunk_index = 0
            filtered_violations: list[str] = []

            def safe_sentence(raw_sentence: str) -> str:
                candidate = RAGRetrievalTool.sanitize_clinical_answer(raw_sentence)
                violations = RAGRetrievalTool.clinical_output_violations(candidate)
                if violations:
                    filtered_violations.extend(violations)
                    candidate = RAGRetrievalTool.remove_unsafe_clinical_sentences(candidate)
                if not candidate or not RAGRetrievalTool.clinical_output_is_safe(candidate):
                    return ""
                return candidate

            for token in client.stream(messages, temperature=0.2) or []:
                sentence_buffer += token
                while True:
                    boundary = re.search(r"[。！？!?；;\n]", sentence_buffer)
                    if not boundary:
                        break
                    raw_sentence = sentence_buffer[:boundary.end()]
                    sentence_buffer = sentence_buffer[boundary.end():]
                    chunk = safe_sentence(raw_sentence)
                    if not chunk:
                        continue
                    llm_generated = True
                    streamed.append(chunk)
                    yield f"data: {json.dumps({'type': 'chunk', 'index': chunk_index, 'content': chunk, 'sentence': True}, ensure_ascii=False)}\n\n"
                    chunk_index += 1

            if sentence_buffer.strip():
                chunk = safe_sentence(sentence_buffer)
                if chunk:
                    llm_generated = True
                    streamed.append(chunk)
                    yield f"data: {json.dumps({'type': 'chunk', 'index': chunk_index, 'content': chunk, 'sentence': True}, ensure_ascii=False)}\n\n"

            generation_mode = "llm-clinical-safe-sentence-stream"
            generation_usage = {"streamed_sentences": len(streamed)}
            if filtered_violations:
                generation_usage["safety_filter"] = list(dict.fromkeys(filtered_violations))[:6]
        elif client.available and can_generate:
            for index, token in enumerate(client.stream(messages, temperature=0.2) or []):
                llm_generated = True
                streamed.append(token)
                yield f"data: {json.dumps({'type': 'chunk', 'index': index, 'content': token}, ensure_ascii=False)}\n\n"
        if not streamed:
            for index, chunk in enumerate(part for part in re.split(r"(?<=[。！？])", fallback_text) if part):
                streamed.append(chunk)
                yield f"data: {json.dumps({'type': 'chunk', 'index': index, 'content': chunk}, ensure_ascii=False)}\n\n"
        text = "".join(streamed)
        result["answer"] = text
        if llm_generated:
            result["generation"] = {"mode": generation_mode or "llm-token-stream",
                                    "provider": client.provider, "model": client.model, "calls": 1,
                                    "usage": generation_usage}
        elif is_retrieval_only:
            result["generation"] = {"mode": "local-stream-fallback", "provider": client.provider,
                                    "model": client.model, "calls": 0}
        result["answer_mode"] = answerMode
        if session_id:
            ConversationService().replace_last_answer(session_id, text, result)
        yield f"data: {json.dumps({'type': 'result', 'content': text, 'result': result}, ensure_ascii=False)}\n\n"
        yield "data: [DONE]\n\n"

    return StreamingResponse(events(), media_type="text/event-stream", headers={"Cache-Control": "no-cache"})


@router.get("/chat/ask/stream/legacy")
def ask_chat_stream(question: str, historyId: str | None = None) -> StreamingResponse:
    result = ask_chat(ChatAskRequest(question=question, historyId=historyId))
    text = result.get("answer", "")

    def events():
        import re
        chunks = [part for part in re.split(r"(?<=[。！？!?])", text) if part]
        for index, chunk in enumerate(chunks):
            yield f"data: {json.dumps({'type': 'chunk', 'index': index, 'content': chunk}, ensure_ascii=False)}\n\n"
        yield f"data: {json.dumps({'type': 'result', 'content': text, 'result': result}, ensure_ascii=False)}\n\n"
        yield "data: [DONE]\n\n"

    return StreamingResponse(events(), media_type="text/event-stream")


@router.get("/chat/history")
def list_chat_history(page: int = 1, pageSize: int = 20) -> dict[str, Any]:
    return ok(ConversationService().list(page, pageSize))


@router.get("/chat/history/{history_id}")
def get_chat_history(history_id: str) -> dict[str, Any]:
    session = ConversationService().load(history_id)
    return ok({"id": session["id"], "title": session["title"], "messages": session["turns"],
               "status": session["status"], "pending_questions": session["pending_questions"]})


@router.post("/chat/history/save")
def save_chat_history(data: dict[str, Any]) -> dict[str, Any]:
    history_id = str(data.get("id") or int(time.time() * 1000))
    session = ConversationService().load(history_id)
    return ok({"id": session["id"], "title": session["title"], "messages": session["turns"]})


@router.delete("/chat/history/{history_id}")
def delete_chat_history(history_id: str) -> dict[str, Any]:
    return ok({"deleted": ConversationService().delete(history_id)})


@router.post("/chat/favorite/{history_id}")
def favorite_chat(history_id: str) -> dict[str, Any]:
    """收藏状态由前端本地维护；接口用于保持新版前端调用契约完整。"""
    session = ConversationService().load(history_id)
    return ok({"id": session["id"], "favorite": True})


def literature_item(document: dict[str, Any]) -> dict[str, Any]:
    return {
        **document,
        "id": str(document.get("id", "")),
        "name": document.get("title", ""),
        "type": document.get("category", "未分类"),
    }


@router.get("/literature/entity/{entity_id}")
def literature_by_entity(entity_id: str) -> dict[str, Any]:
    graph = get_local_graphrag_service()
    entity = graph.entities_by_id.get(entity_id)
    terms = [entity_id]
    if entity:
        terms.extend([entity.name, entity.alias])
    documents = DocumentService().list_documents(limit=100, offset=0)
    items = [
        literature_item(document)
        for document in documents
        if any(term and term in f"{document.get('title', '')}\n{document.get('content', '')}" for term in terms)
    ]
    return ok({"list": items, "total": len(items)})


@router.get("/literature/search")
def search_literature(keyword: str = "") -> dict[str, Any]:
    documents = DocumentService().list_documents(limit=100, offset=0)
    items = [
        literature_item(document)
        for document in documents
        if not keyword or keyword in f"{document.get('title', '')}\n{document.get('source', '')}\n{document.get('content', '')}"
    ]
    return ok({"list": items, "total": len(items)})


@router.get("/literature/{document_id}")
def literature_detail(document_id: int) -> dict[str, Any]:
    document = DocumentService().get_document(document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="literature not found")
    return ok(literature_item(document))


@router.get("/graph/overview")
def graph_overview() -> dict[str, Any]:
    """返回真实全量统计和路径查询所需的全部实体轻量选项。"""
    service = get_local_graphrag_service()
    return {
        "nodeCount": len(service.entities),
        "relationCount": len(service.relations),
        "entities": [
            {"id": entity.id, "label": entity.name, "type": entity.type}
            for entity in service.entities
        ],
    }


@router.get("/graph/full")
def full_graph(limit: int = Query(300, ge=20, le=500)) -> dict[str, Any]:
    service = get_local_graphrag_service()
    # 画布只展示一个关系密集且类型均衡的子图；全量统计与路径选项由
    # /graph/overview 单独提供，避免数千节点进入力导向布局。
    visible_types = {"药材", "方剂", "症状", "证候"}
    degree: dict[str, int] = {}
    for relation in service.relations:
        degree[relation.source_id] = degree.get(relation.source_id, 0) + 1
        degree[relation.target_id] = degree.get(relation.target_id, 0) + 1

    ratios = {"药材": 0.4, "方剂": 0.2, "症状": 0.2, "证候": 0.2}
    entities = []
    selected_ids: set[str] = set()
    for entity_type, ratio in ratios.items():
        quota = max(1, int(limit * ratio))
        candidates = sorted(
            (entity for entity in service.entities if entity.type == entity_type),
            key=lambda entity: (-degree.get(entity.id, 0), entity.id),
        )
        for entity in candidates[:quota]:
            entities.append(entity)
            selected_ids.add(entity.id)

    if len(entities) < limit:
        remaining = sorted(
            (
                entity
                for entity in service.entities
                if entity.type in visible_types and entity.id not in selected_ids
            ),
            key=lambda entity: (-degree.get(entity.id, 0), entity.id),
        )
        entities.extend(remaining[: limit - len(entities)])

    ids = {entity.id for entity in entities}
    nodes = [{"id": e.id, "label": e.name, "type": e.type, "properties": e.properties or {}} for e in entities]
    edges = [
        {"source": r.source_id, "target": r.target_id, "label": r.relation, "properties": {"evidence": r.evidence}}
        for r in service.relations
        if r.source_id in ids and r.target_id in ids
    ]
    return graph_response(nodes, edges)


@router.get("/graph/search")
def search_graph(keyword: str = "", types: str | None = None) -> dict[str, Any]:
    service = get_local_graphrag_service()
    type_set = set(str(types or "").split(",")) if types else set()
    matches = [
        entity
        for entity in service.entities
        if (not type_set or entity.type in type_set)
        and (not keyword or keyword in entity.name or keyword in entity.alias or keyword in entity.description)
    ][:120]
    ids = {entity.id for entity in matches}
    related_edges = [
        relation
        for relation in service.relations
        if relation.source_id in ids or relation.target_id in ids
    ][:300]
    for relation in related_edges:
        ids.add(relation.source_id)
        ids.add(relation.target_id)
    nodes = [
        {"id": entity.id, "label": entity.name, "type": entity.type, "properties": entity.properties or {}}
        for entity in service.entities
        if entity.id in ids
    ]
    edges = [
        {"source": r.source_id, "target": r.target_id, "label": r.relation, "properties": {"evidence": r.evidence}}
        for r in related_edges
    ]
    return graph_response(nodes, edges)


@router.get("/graph/entity/{entity_id}")
def entity_detail(entity_id: str) -> dict[str, Any]:
    service = get_local_graphrag_service()
    entity = service.entities_by_id.get(entity_id)
    if not entity:
        raise HTTPException(status_code=404, detail="entity not found")
    return ok(entity_to_admin(entity, build_relation_property_index(service)))


@router.get("/graph/related/{entity_id}")
def related_graph(entity_id: str, depth: int = 2) -> dict[str, Any]:
    service = get_local_graphrag_service()
    entity = service.entities_by_id.get(entity_id)
    if not entity:
        return graph_response([], [])
    graph = service._expand_graph([entity], max_hops=depth)
    return graph


@router.get("/graph/path")
def graph_path(sourceId: str, targetId: str) -> dict[str, Any]:
    service = get_local_graphrag_service()
    source = service.entities_by_id.get(sourceId)
    target = service.entities_by_id.get(targetId)
    if not source or not target:
        return graph_response([], [])
    graph = service._expand_graph([source, target], max_hops=3)
    return graph


def list_entities_by_type(entity_type: str, page: int = 1, pageSize: int = 20, name: str | None = None) -> dict[str, Any]:
    service = get_local_graphrag_service()
    entities = [
        entity for entity in service.entities
        if entity.type == entity_type and (not name or name in entity.name)
    ]
    start = (page - 1) * pageSize
    relation_properties = build_relation_property_index(service)
    items = [
        entity_to_admin(entity, relation_properties)
        for entity in entities[start : start + pageSize]
    ]
    return ok({"list": items, "total": len(entities)})


def create_entity(kind: str, payload: dict[str, Any]) -> dict[str, Any]:
    entity_type, prefix = ENTITY_TYPE_BY_ADMIN_KIND[kind]
    item = entity_payload_to_item(payload, entity_type, prefix)
    if not item["name"]:
        raise HTTPException(status_code=400, detail="name is required")
    entities = [entity for entity in read_admin_entities() if entity.get("id") != item["id"]]
    entities.append(item)

    overlay = read_admin_overlay()
    overlay["deleted_entity_ids"] = [entity_id for entity_id in overlay["deleted_entity_ids"] if entity_id != item["id"]]
    write_admin_overlay(overlay)
    write_admin_entities(entities)
    return ok(entity_to_admin(Entity(**item)))


def update_entity(kind: str, entity_id: str, payload: dict[str, Any]) -> dict[str, Any]:
    entity_type, prefix = ENTITY_TYPE_BY_ADMIN_KIND[kind]
    service = get_local_graphrag_service()
    current = service.entities_by_id.get(entity_id)
    base_payload = {
        "id": entity_id,
        "name": current.name if current else "",
        "alias": current.alias if current else "",
        "description": current.description if current else "",
        "properties": current.properties or {} if current else {},
    }
    merged = {**base_payload, **payload, "properties": {**base_payload.get("properties", {}), **(payload.get("properties") or {})}}
    item = entity_payload_to_item(merged, entity_type, prefix, entity_id=entity_id)
    if not item["name"]:
        raise HTTPException(status_code=400, detail="name is required")

    entities = [entity for entity in read_admin_entities() if entity.get("id") != entity_id]
    entities.append(item)
    write_admin_entities(entities)
    return ok(entity_to_admin(Entity(**item)))


def delete_entities(ids: list[str]) -> dict[str, Any]:
    id_set = {str(item) for item in ids if str(item)}
    if not id_set:
        return ok({"deleted": 0})

    admin_entities = [entity for entity in read_admin_entities() if entity.get("id") not in id_set]
    write_admin_entities(admin_entities)

    admin_relations = [
        relation
        for relation in read_admin_relations()
        if relation.get("source_id") not in id_set and relation.get("target_id") not in id_set
    ]
    write_admin_relations(admin_relations)

    overlay = read_admin_overlay()
    deleted_entity_ids = set(overlay["deleted_entity_ids"])
    deleted_entity_ids.update(id_set)
    overlay["deleted_entity_ids"] = sorted(deleted_entity_ids)
    write_admin_overlay(overlay)
    return ok({"deleted": len(id_set)})


async def import_entities(kind: str, file: UploadFile) -> dict[str, Any]:
    entity_type, prefix = ENTITY_TYPE_BY_ADMIN_KIND[kind]
    content = await file.read()
    filename = file.filename or ""
    rows: list[dict[str, Any]] = []

    try:
        if filename.lower().endswith(".xlsx"):
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            worksheet = workbook.active
            values = worksheet.iter_rows(values_only=True)
            raw_headers = next(values, ())
            headers = [str(value).strip() if value is not None else "" for value in raw_headers]
            rows = [
                {header: value for header, value in zip(headers, row) if header}
                for row in values
                if any(value not in (None, "") for value in row)
            ]
            workbook.close()
        elif filename.lower().endswith(".json"):
            data = json.loads(content.decode("utf-8-sig"))
            if isinstance(data, dict):
                rows = data.get("entities") or data.get("list") or []
            elif isinstance(data, list):
                rows = data
        elif filename.lower().endswith(".csv"):
            text = content.decode("utf-8-sig")
            rows = list(csv.DictReader(io.StringIO(text)))
        else:
            raise HTTPException(status_code=400, detail="仅支持 .xlsx、.csv 或 .json 文件")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"无法解析导入文件：{exc}") from exc

    existing = {entity.get("id"): entity for entity in read_admin_entities()}
    imported = 0
    errors: list[dict[str, Any]] = []
    imported_ids: set[str] = set()
    for row_number, row in enumerate(rows, start=2):
        if not isinstance(row, dict):
            continue
        payload = {
            "id": row.get("id"),
            "name": row.get("name") or row.get("名称"),
            "alias": row.get("alias") or row.get("别名"),
            "description": row.get("description") or row.get("描述") or "",
            "category": row.get("category") or row.get("分类") or entity_type,
            "properties": {k: v for k, v in row.items() if k not in {"id", "name", "名称", "alias", "别名", "description", "描述"}},
        }
        item = entity_payload_to_item(payload, entity_type, prefix)
        if not item["name"]:
            errors.append({"row": row_number, "field": "name", "message": "名称不能为空"})
            continue
        existing[item["id"]] = item
        imported_ids.add(item["id"])
        imported += 1
    write_admin_entities(list(existing.values()))

    if imported_ids:
        overlay = read_admin_overlay()
        overlay["deleted_entity_ids"] = [
            entity_id for entity_id in overlay["deleted_entity_ids"] if entity_id not in imported_ids
        ]
        write_admin_overlay(overlay)

    return ok({
        "imported": imported,
        "total": len(rows),
        "successCount": imported,
        "errorCount": len(errors),
        "errors": errors,
    })


def export_entities(kind: str) -> Response:
    entity_type, _prefix = ENTITY_TYPE_BY_ADMIN_KIND[kind]
    service = get_local_graphrag_service()
    relation_properties = build_relation_property_index(service)
    rows = [
        entity_to_admin(entity, relation_properties)
        for entity in service.entities if entity.type == entity_type
    ]
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=["id", "name", "type", "alias", "category", "description"])
    writer.writeheader()
    for row in rows:
        writer.writerow({key: row.get(key, "") for key in writer.fieldnames})
    return Response(
        content=output.getvalue().encode("utf-8-sig"),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{kind}.csv"'},
    )


def template_entities(kind: str) -> Response:
    fields_by_kind = {
        "herbs": [
            "id", "name", "alias", "category", "description", "nature_and_flavor",
            "channel_tropism", "efficacy", "indications", "usage_dosage",
            "contraindications", "processing_method",
        ],
        "prescriptions": [
            "id", "name", "alias", "category", "description", "composition", "functions",
            "indications", "preparation", "usage_dosage", "contraindications", "modern_application",
        ],
        "symptoms": [
            "id", "name", "alias", "category", "description", "body_part", "nature",
            "differentiation", "associated_symptoms",
        ],
        "syndromes": [
            "id", "name", "alias", "category", "description", "clinical_manifestations",
            "commonly_used_formulas", "pathogenesis", "treatment_principle",
        ],
    }
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "import_template"
    fields = fields_by_kind[kind]
    worksheet.append(fields)
    worksheet.append(["", "示例名称", "", "", "示例描述"] + [""] * (len(fields) - 5))
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=1, column=len(fields)).column_letter}2"
    for column in worksheet.columns:
        worksheet.column_dimensions[column[0].column_letter].width = max(14, len(str(column[0].value)) + 2)

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{kind}_template.xlsx"'},
    )


@router.get("/admin/herbs")
def admin_herbs(page: int = 1, pageSize: int = 20, name: str | None = None) -> dict[str, Any]:
    return list_entities_by_type("药材", page, pageSize, name)


@router.post("/admin/herbs")
def create_herb(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    return create_entity("herbs", payload)


@router.put("/admin/herbs/{entity_id}")
def update_herb(entity_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    return update_entity("herbs", entity_id, payload)


@router.delete("/admin/herbs/{entity_id}")
def delete_herb(entity_id: str) -> dict[str, Any]:
    return delete_entities([entity_id])


@router.delete("/admin/herbs/batch")
def delete_herbs_batch(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    return delete_entities(payload.get("ids", []))


@router.post("/admin/herbs/import")
async def import_herbs(file: UploadFile = File(...)) -> dict[str, Any]:
    return await import_entities("herbs", file)


@router.get("/admin/herbs/export")
def export_herbs() -> Response:
    return export_entities("herbs")


@router.get("/admin/herbs/template")
def template_herbs() -> Response:
    return template_entities("herbs")


@router.get("/admin/prescriptions")
def admin_prescriptions(page: int = 1, pageSize: int = 20, name: str | None = None) -> dict[str, Any]:
    return list_entities_by_type("方剂", page, pageSize, name)


@router.post("/admin/prescriptions")
def create_prescription(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    return create_entity("prescriptions", payload)


@router.put("/admin/prescriptions/{entity_id}")
def update_prescription(entity_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    return update_entity("prescriptions", entity_id, payload)


@router.delete("/admin/prescriptions/{entity_id}")
def delete_prescription(entity_id: str) -> dict[str, Any]:
    return delete_entities([entity_id])


@router.delete("/admin/prescriptions/batch")
def delete_prescriptions_batch(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    return delete_entities(payload.get("ids", []))


@router.post("/admin/prescriptions/import")
async def import_prescriptions(file: UploadFile = File(...)) -> dict[str, Any]:
    return await import_entities("prescriptions", file)


@router.get("/admin/prescriptions/export")
def export_prescriptions() -> Response:
    return export_entities("prescriptions")


@router.get("/admin/prescriptions/template")
def template_prescriptions() -> Response:
    return template_entities("prescriptions")


@router.get("/admin/symptoms")
def admin_symptoms(page: int = 1, pageSize: int = 20, name: str | None = None) -> dict[str, Any]:
    return list_entities_by_type("症状", page, pageSize, name)


@router.post("/admin/symptoms")
def create_symptom(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    return create_entity("symptoms", payload)


@router.put("/admin/symptoms/{entity_id}")
def update_symptom(entity_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    return update_entity("symptoms", entity_id, payload)


@router.delete("/admin/symptoms/{entity_id}")
def delete_symptom(entity_id: str) -> dict[str, Any]:
    return delete_entities([entity_id])


@router.delete("/admin/symptoms/batch")
def delete_symptoms_batch(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    return delete_entities(payload.get("ids", []))


@router.post("/admin/symptoms/import")
async def import_symptoms(file: UploadFile = File(...)) -> dict[str, Any]:
    return await import_entities("symptoms", file)


@router.get("/admin/symptoms/export")
def export_symptoms() -> Response:
    return export_entities("symptoms")


@router.get("/admin/symptoms/template")
def template_symptoms() -> Response:
    return template_entities("symptoms")


@router.get("/admin/syndromes")
def admin_syndromes(page: int = 1, pageSize: int = 20, name: str | None = None) -> dict[str, Any]:
    return list_entities_by_type("证候", page, pageSize, name)


@router.post("/admin/syndromes")
def create_syndrome(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    return create_entity("syndromes", payload)


@router.put("/admin/syndromes/{entity_id}")
def update_syndrome(entity_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    return update_entity("syndromes", entity_id, payload)


@router.delete("/admin/syndromes/{entity_id}")
def delete_syndrome(entity_id: str) -> dict[str, Any]:
    return delete_entities([entity_id])


@router.delete("/admin/syndromes/batch")
def delete_syndromes_batch(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    return delete_entities(payload.get("ids", []))


@router.post("/admin/syndromes/import")
async def import_syndromes(file: UploadFile = File(...)) -> dict[str, Any]:
    return await import_entities("syndromes", file)


@router.get("/admin/syndromes/export")
def export_syndromes() -> Response:
    return export_entities("syndromes")


@router.get("/admin/syndromes/template")
def template_syndromes() -> Response:
    return template_entities("syndromes")


@router.get("/admin/relations")
def admin_relations(
    page: int = 1,
    pageSize: int = 20,
    sourceName: str | None = None,
    targetName: str | None = None,
    relation: str | None = None,
) -> dict[str, Any]:
    service = get_local_graphrag_service()
    items = [
        relation_to_admin_item(r)
        for r in service.relations
    ]
    if sourceName:
        items = [item for item in items if sourceName in item.get("source_name", "")]
    if targetName:
        items = [item for item in items if targetName in item.get("target_name", "")]
    if relation:
        items = [item for item in items if relation == item.get("relation")]
    start = (page - 1) * pageSize
    return ok({"list": items[start : start + pageSize], "total": len(items)})


@router.post("/admin/relations")
def create_relation(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    source_id = str(payload.get("source_id") or "").strip()
    target_id = str(payload.get("target_id") or "").strip()
    relation = str(payload.get("relation") or "").strip()
    if not source_id or not target_id or not relation:
        raise HTTPException(status_code=400, detail="source_id, target_id and relation are required")

    service = get_local_graphrag_service()
    source = service.entities_by_id.get(source_id)
    target = service.entities_by_id.get(target_id)
    item = {
        "source_id": source_id,
        "source_name": str(payload.get("source_name") or (source.name if source else "")),
        "source_type": str(payload.get("source_type") or (source.type if source else "")),
        "relation": relation,
        "target_id": target_id,
        "target_name": str(payload.get("target_name") or (target.name if target else "")),
        "target_type": str(payload.get("target_type") or (target.type if target else "")),
        "evidence": str(payload.get("evidence") or "后台管理录入"),
    }
    relation_id = relation_key_from_parts(source_id, relation, target_id)
    items = [row for row in read_admin_relations() if relation_key_from_parts(str(row.get("source_id")), str(row.get("relation")), str(row.get("target_id"))) != relation_id]
    items.append(item)

    overlay = read_admin_overlay()
    overlay["deleted_relation_keys"] = [key for key in overlay["deleted_relation_keys"] if key != relation_id]
    write_admin_overlay(overlay)
    write_admin_relations(items)
    return ok(relation_to_admin_item(item))


def delete_relations(ids: list[str]) -> dict[str, Any]:
    key_set = {str(item) for item in ids if str(item)}
    if not key_set:
        return ok({"deleted": 0})

    items = [
        row
        for row in read_admin_relations()
        if relation_key_from_parts(str(row.get("source_id")), str(row.get("relation")), str(row.get("target_id"))) not in key_set
    ]
    write_admin_relations(items)
    overlay = read_admin_overlay()
    deleted_relation_keys = set(overlay["deleted_relation_keys"])
    deleted_relation_keys.update(key_set)
    overlay["deleted_relation_keys"] = sorted(deleted_relation_keys)
    write_admin_overlay(overlay)
    return ok({"deleted": len(key_set)})


@router.delete("/admin/relations/{relation_id}")
def delete_relation(relation_id: str) -> dict[str, Any]:
    return delete_relations([relation_id])


@router.delete("/admin/relations/batch")
def delete_relations_batch(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    return delete_relations(payload.get("ids", []))


@router.get("/stats/platform")
def platform_stats() -> dict[str, Any]:
    service = get_local_graphrag_service()
    counts: dict[str, int] = {}
    for entity in service.entities:
        counts[entity.type] = counts.get(entity.type, 0) + 1
    return {
        "totalHerbs": counts.get("药材", 0),
        "totalPrescriptions": counts.get("方剂", 0),
        "totalSymptoms": counts.get("症状", 0),
        "totalSyndromes": counts.get("证候", 0),
        "totalRelations": len(service.relations),
        "totalQuestions": len(HistoryService().list_history(limit=100, offset=0)),
        "totalCases": TeachingCaseService().list(1, 1)["total"],
    }


@router.get("/stats/categories")
def category_stats() -> dict[str, Any]:
    service = get_local_graphrag_service()
    counts: dict[str, int] = {}
    for entity in service.entities:
        counts[entity.type] = counts.get(entity.type, 0) + 1
    return counts


@router.get("/stats/popular/{entity_type}")
def popular_entities(entity_type: str, limit: int = 10) -> list[dict[str, Any]]:
    mapping = {"herbs": "药材", "prescriptions": "方剂", "symptoms": "症状"}
    service = get_local_graphrag_service()
    target_type = mapping.get(entity_type, entity_type)
    return [
        {"name": entity.name, "value": index + 1, "id": entity.id}
        for index, entity in enumerate([e for e in service.entities if e.type == target_type][:limit])
    ]


@router.get("/stats/daily-questions")
def daily_questions(days: int = 7) -> list[dict[str, Any]]:
    return [{"date": now_iso()[:10], "count": 0} for _ in range(days)]
