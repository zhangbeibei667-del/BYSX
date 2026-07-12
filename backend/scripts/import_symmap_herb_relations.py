from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from openpyxl import load_workbook


CONFIG = {
    "TCM_symptom": {"id": "TCM symptom id", "name": "TCM symptom name", "type": "症状", "relation": "关联症状"},
    "Syndrome": {"id": "Syndrome id", "name": "Syndrome name", "type": "证候", "relation": "关联证候"},
    "MM_symptom": {"id": "MM symptom id", "name": "MM symptom name", "type": "现代症状", "relation": "关联现代症状"},
    "Mol": {"id": "Ingredient id", "name": "Ingredient name", "type": "成分", "relation": "包含成分"},
}


def clean(value) -> str:
    text = str(value or "").strip().strip("\ufeff")
    return "" if text.lower() in {"none", "null", "nan"} else text


def load_entities(directory: Path) -> tuple[dict, dict, dict]:
    entities, herbs, herbs_by_name = {}, {}, {}
    for path in directory.glob("*.json"):
        try:
            data = json.loads(path.read_text(encoding="utf-8-sig"))
        except Exception:
            continue
        data = data.get("entities", []) if isinstance(data, dict) else data
        for item in data if isinstance(data, list) else []:
            entity_id = clean(item.get("id"))
            if not entity_id:
                continue
            entities[entity_id] = item
            props = item.get("properties") or {}
            if item.get("type") == "药材" or props.get("symmap_component") == "SMHB" or entity_id.startswith("SMHB"):
                herbs[entity_id] = clean(item.get("name"))
                herbs_by_name[clean(item.get("name"))] = item
    return entities, herbs, herbs_by_name


def import_relations(raw_dir: Path, graph_data_dir: Path, herb_workbook: Path | None = None) -> dict:
    entity_dir, relation_dir = graph_data_dir / "entities", graph_data_dir / "relations"
    relation_dir.mkdir(parents=True, exist_ok=True)
    entities, herb_names, herbs_by_name = load_entities(entity_dir)
    symmap_names = {}
    if herb_workbook and herb_workbook.exists():
        rows = load_workbook(herb_workbook, read_only=True, data_only=True).active.iter_rows(min_row=2, values_only=True)
        symmap_names = {f"SMHB{int(row[0]):05d}": clean(row[1]) for row in rows if isinstance(row[0], (int, float))}
    new_entities, relations = {}, {}
    files = rows = 0
    for component, config in CONFIG.items():
        for path in sorted((raw_dir / component).glob("SMHB*.csv")) if (raw_dir / component).exists() else []:
            files += 1
            symmap_herb_id = path.stem
            herb_name = herb_names.get(symmap_herb_id) or symmap_names.get(symmap_herb_id) or symmap_herb_id
            canonical = herbs_by_name.get(herb_name)
            herb_id = canonical["id"] if canonical else symmap_herb_id
            if herb_id not in entities and herb_id not in new_entities:
                new_entities[herb_id] = {"id": herb_id, "name": herb_name, "type": "药材", "alias": "",
                                         "description": "", "properties": {"source": "SymMap v2 related_components",
                                                                            "symmap_component": "SMHB"}}
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                for row in csv.DictReader(handle):
                    rows += 1
                    target_id, target_name = clean(row.get(config["id"])), clean(row.get(config["name"]))
                    if not target_id or not target_name:
                        continue
                    if target_id not in entities:
                        new_entities[target_id] = {"id": target_id, "name": target_name, "type": config["type"],
                                                   "alias": "", "description": "",
                                                   "properties": {"source": "SymMap v2 related_components",
                                                                  "symmap_component": component}}
                    key = (herb_id, config["relation"], target_id)
                    relations[key] = {"source_id": herb_id, "source_name": herb_name,
                                      "relation": config["relation"], "target_id": target_id,
                                      "target_name": target_name,
                                      "evidence": f"SymMap v2 related_components: {herb_id} -> {component}"}
    entity_output = entity_dir / "entities_symmap_relation_targets.json"
    relation_output = relation_dir / "relations_symmap_herb.json"
    entity_output.write_text(json.dumps(list(new_entities.values()), ensure_ascii=False, indent=2), encoding="utf-8")
    relation_output.write_text(json.dumps(list(relations.values()), ensure_ascii=False, indent=2), encoding="utf-8")
    return {"files": files, "rows": rows, "relations": len(relations), "new_entities": len(new_entities),
            "entity_output": str(entity_output), "relation_output": str(relation_output)}


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--raw-dir", default="integrated_entities_graphrag/external_data/symmap_v2/herb_relations/raw")
    parser.add_argument("--graph-data-dir", default="integrated_entities_graphrag/data")
    parser.add_argument("--herb-workbook", default="SymMap v2.0, SMHB file.xlsx")
    args = parser.parse_args()
    print(json.dumps(import_relations(Path(args.raw_dir), Path(args.graph_data_dir), Path(args.herb_workbook)), ensure_ascii=False, indent=2))
